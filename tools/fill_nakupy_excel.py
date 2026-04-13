#!/usr/bin/env python3
"""Doplní prázdný sloupec produkt (kopie střelnice / kategorie) a sjednotí názvy střelnic v listu nákupy."""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

STREL_CANON = [
    "střelnice Břidličná",
    "střelnice Corrado Ostrava",
    "střelnice Krnov",
    "střelnice Lazce Olomouc",
    "střelnice Polárka FM",
    "střelnice Třinec",
    "střelnice Uherský Brod",
]


def norm_strelnice(raw: object) -> str:
    s = str(raw or "").strip()
    if not s:
        return s
    s = s.replace("Střelnie", "střelnice").replace("střelnie", "střelnice")
    if s.lower().startswith("střeln"):
        key = s.lower()
        for c in STREL_CANON:
            if c.lower() == key:
                return c
    return s


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: fill_nakupy_excel.py <path.xlsx>", file=sys.stderr)
        return 2
    path = Path(sys.argv[1]).expanduser().resolve()
    if not path.is_file():
        print(f"Missing file: {path}", file=sys.stderr)
        return 1

    wb = load_workbook(path, data_only=False)
    ws = wb.active
    headers = [(ws.cell(1, c).value, c) for c in range(1, ws.max_column + 1)]
    col = {}
    for h, idx in headers:
        if h is None:
            continue
        col[str(h).strip().lower()] = idx

    k_cat = col.get("kategorie")
    k_prod = col.get("produkt")
    k_str = col.get("střelnice") or col.get("strelnice")
    if not k_prod:
        print("Chybí sloupec produkt v prvním řádku.", file=sys.stderr)
        return 1

    filled = 0
    norm_s = 0
    for r in range(2, ws.max_row + 1):
        prod = ws.cell(r, k_prod).value if k_prod else None
        st = ws.cell(r, k_str).value if k_str else None
        kat = ws.cell(r, k_cat).value if k_cat else None
        if k_str and st not in (None, ""):
            new_s = norm_strelnice(st)
            if new_s != str(st).strip():
                ws.cell(r, k_str, new_s)
                norm_s += 1
        if prod is None or (isinstance(prod, str) and not prod.strip()):
            if st not in (None, "") and str(st).strip():
                ws.cell(r, k_prod, str(st).strip())
                filled += 1
            elif kat not in (None, "") and str(kat).strip():
                ws.cell(r, k_prod, str(kat).strip())
                filled += 1

    wb.save(path)
    print(f"OK: {path} — doplněno {filled} buněk produkt, upraveno {norm_s} buněk střelnice")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
