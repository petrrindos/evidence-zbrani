#!/usr/bin/env python3
"""Normalize ZBRANĚ.xlsx for import into evidence-zbrani.html (střelnice názvy, překlepy)."""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

STREL_CANON = [
    "střelnice Břidličná",
    "střelnice Corrado Ostrava",
    "střelnice Krnov",
    "střelnice Lazce Olomouc",
    "střelnice Polárka FM",
    "střelnice Třinec",
    "střelnice Uherský Brod",
]


def norm_strelnice(raw: object) -> str:
    s = str(raw or "").strip()
    if not s:
        return s
    s = s.replace("Střelnie", "střelnice").replace("střelnie", "střelnice")
    if not s.lower().startswith("střelnice"):
        # např. "Střelnice Polárka FM"
        if s.lower().startswith("střeln"):
            pass
        else:
            return s
    key = s.lower()
    for c in STREL_CANON:
        if c.lower() == key:
            return c
    return s


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: normalize_zbrane_xlsx.py <path.xlsx>", file=sys.stderr)
        return 2
    path = Path(sys.argv[1]).expanduser().resolve()
    if not path.is_file():
        print(f"Missing file: {path}", file=sys.stderr)
        return 1

    wb = load_workbook(path, data_only=False)
    ws = wb.active
    # Row 1: headers datum | střelnice | zbraň | náboje | počet nábojů
    headers = [ws.cell(1, c).value for c in range(1, 6)]
    h0 = str(headers[0] or "").strip().lower() if headers[0] else ""
    if h0 != "datum":
        print(f"Warning: A1 expected 'datum', got {headers[0]!r}", file=sys.stderr)

    changed = 0
    for r in range(2, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        if a is None and b is None and ws.cell(r, 3).value is None:
            continue
        new_b = norm_strelnice(b)
        if new_b != (str(b).strip() if b is not None else ""):
            ws.cell(r, 2, new_b)
            changed += 1

    wb.save(path)
    print(f"OK: {path} (upraveno {changed} buněk ve sloupci střelnice)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
